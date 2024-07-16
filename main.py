# Selenium Modules
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium_stealth import stealth


from bs4 import BeautifulSoup

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

import os
import io

import openpyxl

#Dropbox modules
import dropbox
from dropbox.exceptions import AuthError, ApiError
from dropbox.common import PathRoot
from dropbox import DropboxOAuth2FlowNoRedirect
from dropbox.dropbox_client import DropboxTeam

#Email modules
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Printing helpers
from color_printer import printb,printg,printr  

############## SCRAPING & WEB BROWSING FNS ###################################################

# Fintel Login through Selenium function
def login_to_fintel(username, password):   
    """ Sets up Web Driver and completes authentication through Fintel. 
    Returns WebDriver for user in later scraping fns."""
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless=new')
    # chrome_options.add_argument("start-maximized")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=chrome_options)

    stealth(driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
        )
    
    # First login to get access to all the data

    driver.get("https://fintel.io/auth/login")

    # driver.get_screenshot_as_file("screenshot.png")

    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CLASS_NAME,"btn-primary"))
    )

    driver.find_element(By.ID,"username").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.CLASS_NAME,"btn-primary").click()

    error_message = "The credentials you entered are not valid"

    errors = driver.find_elements("css selector",".alert-danger")

    if any(error_message in e.text for e in errors):
        printr("[!] Login failed")
    else:
        printg("[+] Login successful")
    
    return driver

# Fintel scraping function
def scrape_fintel_data(driver, stock_ticker, table_name):
    '''Connects to Fintel using Selenium. Uses Beautiful Soup to parse HTML. Converts HTML to a Pandas DataFrame. Returns the DataFrame'''

    url = f"https://fintel.io/ss/us/{stock_ticker}"

    driver.get(url)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, table_name))
    )

    page_source = driver.page_source

    soup = BeautifulSoup(page_source, 'html.parser')

    # Find the table(s) containing the data you need
    table = soup.find('table', {'id': table_name})

    # Extract headers from the first row (assuming they are in <th> elements)
    headers = [th.text.strip().replace("FINRA","") for th in table.find('tr').find_all('th')]
    headers = [header for header in headers if header != ""]

    # Extract data from subsequent rows
    data_rows = []
    for row in table.find_all('tr')[1:]:
        data = [td.text.strip() for td in row.find_all('td') if td.text.strip() not in ["+","/","="]]
        data_rows.append(data)

    # Organize data into a DataFrame
    data = pd.DataFrame(data_rows, columns=headers)

    # Clean and process data

    # Convert Market Dates to date format
    # Based on tables being scraped from Fintel, if Market Date returns an error, there will be a Settlement and Publication Date
    try:
        data['Market Date'] = pd.to_datetime(data['Market Date'], errors= 'coerce').dt.strftime('%Y-%m-%d')
    except:
        data['Settlement Date'] = pd.to_datetime(data['Settlement Date'], errors= 'coerce').dt.strftime('%Y-%m-%d')
        data['Publication Date'] = pd.to_datetime(data['Publication Date'], errors= 'coerce').dt.strftime('%Y-%m-%d')

    # Clean up column headers
    data.columns = [str(col).replace("BX","BX ").replace("Aggregate","Aggregate ").replace("*","").strip() for col in data.columns]

    return data

def create_yahoo_driver():
    """Creates and returns a new WebDriver to be used for Yahoo scraping fns."""
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument("start-maximized")
    # chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    # chrome_options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=chrome_options)
    
    return driver

def scrape_yahoo_data(driver, stock_ticker):
    '''Connects to Yahoo using Selenium. Uses Beautiful Soup to parse HTML. Converts HTML to a Pandas DataFrame. Returns the DataFrame'''

    url = f"https://finance.yahoo.com/quote/{stock_ticker}/history"

    driver.get(url)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "table"))
    )

    page_source = driver.page_source

    soup = BeautifulSoup(page_source, 'html.parser')

    # Find the table(s) containing the data you need
    table = soup.find('table')

    headers = [th.text[:6].replace("*","").strip() for th in table.find('tr').find_all('th')]

    data_rows = []
    for row in table.find_all('tr')[1:11]:
        data = [td.text.strip() for td in row.find_all('td')]
        data_rows.append(data)
    
    data = pd.DataFrame(data_rows,columns=headers)

    data = data[['Date','Close','Volume']] # Get the only columns we want

    data['Date'] = pd.to_datetime(data['Date'],format="%b %d, %Y").dt.strftime('%Y-%m-%d')

    data.iloc[:,1:] = data.iloc[:,1:].map(lambda x: pd.to_numeric(x, errors='ignore'))

    data.rename(columns={'Date': 'Market Date'}, inplace=True)
    data.rename(columns={'Close': 'Close Price'}, inplace=True)

    data.loc[:,'Close Price'] = data.loc[:,'Close Price'].apply(lambda x: f"${x}")

    return data

def scrape_all_data(stock_tickers):
    """ Wrapper function for scrapers from different sources for logic abstraction
        Returns a dictionary of dataframes with all the scraped data organized by source and stock ticker
    """
    data = {}

    # Login to Fintel
    printb("Logging into Fintel...")
    username = "" # Removed for demo
    password = "" # Removed for demo
    fintel_driver = login_to_fintel(username, password)

    for stock_ticker in stock_tickers:

        printb(f"Scraping Fintel data for {stock_ticker}...")
        data[stock_ticker] = {}

        # Scrape data from Fintel
        finra_table_title = "short-sale-volume-finra-table"
        combined_table_title = "short-sale-volume-combined-table"
        historical_table_title = "short-interest-nasdaq-table"

        printb("Scraping FINRA data...")
        short_interest_finra_data = scrape_fintel_data(fintel_driver, stock_ticker, finra_table_title)
        data[stock_ticker]["finra_data"] = short_interest_finra_data
        printg("[+] Scraped!")

        printb("Scraping Combined data...")
        short_interest_combined_data = scrape_fintel_data(fintel_driver, stock_ticker, combined_table_title)
        data[stock_ticker]["combined_data"] = short_interest_combined_data
        printg("[+] Scraped!")

        printb("Scraping Historical data...")
        short_interest_historical_data = scrape_fintel_data(fintel_driver, stock_ticker, historical_table_title)
        data[stock_ticker]["historical_data"] = short_interest_historical_data
        printg("[+] Scraped!")

    fintel_driver.quit()

    yahoo_driver = create_yahoo_driver()

    for stock_ticker in stock_tickers:

        printb(f"Scraping Yahoo data for {stock_ticker}...")
        yahoo_data = scrape_yahoo_data(yahoo_driver,stock_ticker)
        data[stock_ticker]["yahoo_data"] = yahoo_data
        printg("[+] Scraped!")

        
    yahoo_driver.quit()

    return data

################ EXCEL FNS ################################################################3

# Excel Format helper
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from datetime import datetime
def format_workbook(workbook):
    '''Excel formatting helper for font sizes, colors, etc.'''
    for sheet in workbook.worksheets:
        # format the ticker
        sheet['A1'].font = Font(size=20, bold=True)

        # format the headers
        for cell in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=2, max_row=2):
            for cell_obj in cell: 
                # print(type(cell))
                # print(type(sheet['A1']))
                white = "ffffff"
                cell_obj.font = Font(bold=True,size=15,color=white)
                blue = "2181ff"
                cell_obj.fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")

        # format the numbers
        for cell in sheet.iter_rows(min_row=3,min_col=2):
            for cell_obj in cell: 
                cell_obj.font = Font(size=12)

        #set column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5
            sheet.column_dimensions[column].width = adjusted_width
            
    return workbook
    
# Excel Spreadsheet creation
def create_excel_sheet(data, sheet_name, stock_ticker, current_workbook=None):
    '''Turns a pandas dataframe into an excel sheet then returns the excel workbook
        Takes a data frame, sheet_name, stock ticker and optional current workbook
        Either adds sheet to current_workbook or creates new workbook with the sheet and returns the workbook'''
    
    printb(f"Creating the {sheet_name} sheet...")

    
    # Open existing workbook if it exists, or create a new one
    if current_workbook:
        workbook = current_workbook
        sheet = get_sheet_helper(workbook,sheet_name)
        printg("Retrieved the current sheet!")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        printb("No current sheet. Creating a new sheet.")

    # Clear all rows in the sheet
    # sheet.delete_rows(1, sheet.max_row)

    # Write ticker to top of file
    sheet['A1'] = f'Ticker = {stock_ticker.upper()}'

    # Write headers to Excel spreadsheet if not already there
    if sheet['A2'].value != 'Market Date':
        write_df_row_to_sheet(data.columns,2,sheet)

    # Write data to Excel spreadsheet unless date already is in the sheet
    current_dates = [str(date) for date in [cell.value for cell in sheet['A']][2:]]
    for _, row in data[::-1].iterrows():
        if row['Market Date'] not in current_dates:
            # Insert the row at the top of the sheet
            printg(f"Adding {row['Market Date']} to the sheet...")
            sheet.insert_rows(3) # insert blank row at 3 
            
            write_df_row_to_sheet(row,3,sheet)
        else:
            printb(row['Market Date'], "is already in the sheet.")

    return workbook

# Sheet Check helper for create excel
def get_sheet_helper(workbook, sheet_title):
    'If sheet with title exists returns that sheet. Else returns new sheet.'
    for sheet_name in workbook.sheetnames:
        if sheet_name == sheet_title:
            return workbook[sheet_name]
    return workbook.create_sheet(title=sheet_title)

def convert_excel_to_binary(excel_data):
    # Converts a pandas df dataframe into an in memory BytesIO object so it can be uploaded to Dropbox
    excel_io = io.BytesIO()
    excel_data.save(excel_io)

    excel_binary = excel_io.getvalue()

    return excel_binary

def write_df_row_to_sheet(row, row_num, sheet):
    '''Function that takes in a dataframe row and openpyxl sheet and writes everything in the row to the row_num in the sheet'''
    for col_num, val in (enumerate(row.tolist(),1)):
            sheet.cell(row=row_num, column = col_num, value=val)
    return sheet

############# DROPBOX FNS ############################################################333


def get_refresh_token(APP_KEY, APP_SECRET):
    '''Initial dropbox authentication. Prompts user to log in, authorize and retrieve authorization code. 
    Then retrieves the access code and refresh token for future use. This method should only be run once, 
    then the refresh token (does not expire) can be used in subsequent calls.'''

    auth_flow = DropboxOAuth2FlowNoRedirect(APP_KEY, APP_SECRET, token_access_type='offline')

    authorize_url = auth_flow.start()
    print("1. Go to: " + authorize_url)
    print("2. Click \"Allow\" (you might have to log in first).")
    print("3. Copy the authorization code.")
    auth_code = input("Enter the authorization code here: ").strip()

    try:
        oauth_result = auth_flow.finish(auth_code)
    except Exception as e:
        print('Error: %s' % (e,))
        exit(1)

    with dropbox.Dropbox(oauth2_access_token=oauth_result.access_token) as dbx:
        print("Successfully set up client!")

    print(oauth_result)
    print("access token:", oauth_result.access_token)
    print("refresh token:", oauth_result.refresh_token)    
    return oauth_result.refresh_token

def create_dropbox_instance(APP_KEY, APP_SECRET, REFRESH_TOKEN):
    dbx = DropboxTeam(
                    app_key = APP_KEY,
                    app_secret= APP_SECRET,
                    oauth2_refresh_token= REFRESH_TOKEN
    )
    for member in dbx.team_members_list().members:
        if member.profile.email == "": # email removed for demo
            my_id = member.profile.team_member_id
            return dbx.as_user(my_id)
    return None

def set_correct_path_root(dbx):
    """Sets the path root to the team root instead of the user root to access a team level Dropbox folder"""
    # Make an API call to get current account information
    account_info = dbx.users_get_current_account()

    # Extract the root namespace ID from the response to use team root instead of user root
    root_namespace_id = account_info.root_info.root_namespace_id

    # Instantiate PathRoot object with correct root id
    root_path_root = PathRoot.root(root_namespace_id)

    # Instantiate dropbox instance with the correct path root
    dbx = dbx.with_path_root(root_path_root)

    return dbx

def download_dropbox_file(dbx,file_path):
    """Takes in a dropbox instance and file path for a file and returns a bytes representation of the dropbox object"""
    try:
        _, response = dbx.files_download(file_path)
        printg(f"Downloaded file at {file_path}")
        return response.content
    except:
        printr(f"No dropbox file found at {file_path}")
        return None
    
def upload_excel_to_dropbox(dbx, excel_data, dropbox_folder_path, dropbox_file_name):
    try:

        # Convert Excel to binary
        excel_binary = convert_excel_to_binary(excel_data)
        # Set the correct path root for the dbx object
        # dbx = set_correct_path_root(dbx)
        # Specify the Dropbox path for the file
        dropbox_path = f'{dropbox_folder_path}/{dropbox_file_name}'

        # Upload the Excel file to Dropbox

        dbx.files_upload(excel_binary, dropbox_path, mode=dropbox.files.WriteMode('overwrite'))

        printg(f'[+] Successfully uploaded Excel file to {dropbox_path}')
    except AuthError as e:
        printr(f'Error authenticating with Dropbox: {e}')
    except ApiError as e:
        printr(f'Dropbox API error: {e}')
    except Exception as e:
        printr(f'An unexpected error occurred: {e}')

############ EMAIL FNS ##############################################################3

def send_email(receiver_emails, subject, body):

    port = 465  # For SSL
    email = "" # Removed for demo
    password = ""  # Application password (removed for demo)

    # Create a secure SSL context
    context = ssl.create_default_context()

    sender_email = "" # Removed for demo

    # Create the MIME object
    message = MIMEMultipart()
    message['Subject'] = subject
    message['From'] = "" + f'<{sender_email}>' # Removed for demo
    message['To'] = ", ".join(receiver_emails)

    # Attach HTML body
    html_message = MIMEText(body, 'html')
    message.attach(html_message)

    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.ehlo()
        server.login(email, password)
        server.sendmail(sender_email, receiver_emails, message.as_string())
        server.quit()

########################################################################################

def main():

    # Combine fintel and yahoo data
    try: 
        os.system("taskkill /F /IM excel.exe")
    except:
        print("Tried to close excel but wasn't open")

    # Stocks to scrape data from
    stock_tickers = [""]
    data = scrape_all_data(stock_tickers)

    # Combine dataframes for formatting the email tables and excel sheets

    for ticker in stock_tickers: 

        merged_df = pd.merge(data[ticker]['combined_data'],data[ticker]['yahoo_data'])
        
        # replace empty ASV values with the FINRA volumes
        merged_df['Aggregate Short Volume'] = merged_df['Aggregate Short Volume'].replace('',np.nan)
        merged_df['Aggregate Short Volume'] = merged_df['Aggregate Short Volume'].fillna(merged_df['Short Volume'])
        merged_df = merged_df[["Market Date","Close Price","Aggregate Short Volume","Volume"]]

        # Convert ASV and V columns to numeric to divide them 
        merged_df['Aggregate Short Volume'] = pd.to_numeric(merged_df['Aggregate Short Volume'].str.replace(',',''))
        merged_df['Volume'] = pd.to_numeric(merged_df['Volume'].str.replace(',',''))

        # Create the ASV/V column
        merged_df['ASV/V'] = merged_df['Aggregate Short Volume'] / merged_df['Volume']
        
        # Fill NaNs with empty string
        merged_df = merged_df.fillna('')

        # Format percentages and floats
        merged_df['ASV/V'] = merged_df['ASV/V'].apply(lambda x: f"{x:.2%}" if x != '' else x)
        merged_df['Aggregate Short Volume'] = merged_df['Aggregate Short Volume'].apply(lambda x: f"{x:,.0f}" if x != '' else x)
        merged_df['Volume'] = merged_df['Volume'].apply(lambda x: f"{x:,.0f}" if x != '' else x)
        data[ticker]['merged_data'] = merged_df

    merged_dfs = []
    historical_dfs = []
    for ticker in data.keys():
        merged_dfs.append(data[ticker]["merged_data"])
        historical_dfs.append(data[ticker]["historical_data"])

    # Set the subject, body and recipients of the email
        
    # Set the subject, body and recipients of the email
    today = data[""]["merged_data"].iloc[0,0]
    subject = f"  Short Interest Data for {today}"
    stock_names = {"": "", "": ""}

    body = "<html><body>"
    body += f"""
                <p>
                This is a short interest update for ___stock A___ and __stock B___ as of {datetime.now().strftime("%Y-%m-%d %H:%M")}.<br><br>
                See current short interest data and historical NASDAQ data for each stock in the tables below.
                The data was sourced from <a href="https://finance.yahoo.com/">Yahoo Finance</a> and <a href="https://fintel.io/">Fintel.</a>
                </p>
            """

    for ticker in stock_tickers:
        result = ""
        result += f"<h2>{stock_names[ticker]}</h2>"
        result += "<h3>Short Interest Data</h3>"
        result += data[ticker]["merged_data"].to_html(index=False)
        result += "<br>"
        result += "<h3>Historical Data</h3>"
        result += data[ticker]["historical_data"].to_html(index=False)
        result += "<br><br><br>"

        body += result

    body += "</body></html>"
    
    recipients = [""]

    # Send the email
    send_email(recipients, subject, body)
    printg(f"[+] Email(s) successfully sent to {','.join(recipients)} at {datetime.now()}")

    email_dict = {}

    for ticker in stock_names.keys():
        email_dict[ticker] = {}
        subject = f"{ticker.upper()} Short Interest Data for {today}"
        body = "<html><body>"
        body += f"""
                    <p>
                    This is a short interest update for {stock_names[ticker]} as of {datetime.now().strftime("%Y-%m-%d %H:%M")}.<br><br>
                    See current short interest data and historical NASDAQ data for each stock in the tables below.
                    The data was sourced from <a href="https://finance.yahoo.com/">Yahoo Finance</a> and <a href="https://fintel.io/">Fintel.</a>
                    </p>
                """

        result = ""
        result += f"<h2>{stock_names[ticker]}</h2>"
        result += "<h3>Short Interest Data</h3>"
        result += data[ticker]["merged_data"].to_html(index=False)
        result += "<br>"
        result += "<h3>Historical Data</h3>"
        result += data[ticker]["historical_data"].to_html(index=False)
        result += "<br><br><br>"

        body += result

        body += "</body></html>"

        email_dict[ticker]["subject"] = subject
        email_dict[ticker]["body"] = body

    
    recipients_ = [""] # Removed for demo

    recipients_ = [""] # Removed for demo

    # Send the email
        
    send_email(recipients_, email_dict[""]["subject"], email_dict[""]["body"])
    printg(f"[+] Email(s) successfully sent to {','.join(recipients_)} at {datetime.now()}")

    send_email(recipients_, email_dict[""]["subject"], email_dict[""]["body"])
    printg(f"[+] Email(s) successfully sent to {','.join(recipients_)} at {datetime.now()}")

    # Update and Upload File to Dropbox 

    APP_KEY = "" # Removed for demo
    APP_SECRET = "" # Removed for demo
    
    # Use next line on initial call to get the refresh token then comment out and hardcode refresh token
    # REFRESH_TOKEN = get_refresh_token(APP_KEY,APP_SECRET)
    REFRESH_TOKEN = "" # Removed for demo

    # Create the Dropbox instance
    dbx = create_dropbox_instance(APP_KEY, APP_SECRET, REFRESH_TOKEN)
    for stock_ticker in stock_tickers:
        # Dropbox Authentication and Parameters
        dropbox_folder_path = '/Short Interest -  & /Automated'
        dropbox_file_name = f"{stock_ticker.upper()}_Short_Interest_Data.xlsx"
        dropbox_file_path = f"{dropbox_folder_path}/{dropbox_file_name}"



        # Download the existing sheet from Dropbox in binary form
        dbx_workbook = download_dropbox_file(dbx,dropbox_file_path)

        # Convert the downloaded binary data to an Excel workbook
        if dbx_workbook:
            dbx_workbook = openpyxl.load_workbook(io.BytesIO(dbx_workbook))

        # Clean up data being added to tables 
        finra_data = data[stock_ticker]["finra_data"]
        combined_data = data[stock_ticker]["combined_data"]

        finra_data.iloc[:,1:] = finra_data.iloc[:,1:].map(lambda a: pd.to_numeric(a.replace(",","")))
        combined_data.iloc[:,1:] = combined_data.iloc[:,1:].map(lambda a: pd.to_numeric(a.replace(",","")))
        
        workbook = create_excel_sheet(finra_data, "finra-table", stock_ticker, dbx_workbook)

        workbook = create_excel_sheet(combined_data, "combined-table", stock_ticker, dbx_workbook)

        workbook = format_workbook(workbook)

        # Upload the Excel sheet to Dropbox directly from memory
        upload_excel_to_dropbox(dbx, workbook, dropbox_folder_path, dropbox_file_name)

# Call the main function if the script is run directly
if __name__ == "__main__":
    main()
